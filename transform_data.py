import pandas as pd
import os
from fuzzywuzzy import fuzz
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging
import re
import subprocess

# Configure logging
logging.basicConfig(
    filename='/opt/tables/app.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuration
WORKING_DIR = "/opt/tables"
SKLAD_FILE = os.path.join(WORKING_DIR, "sklad.xlsx")
REESTR_FILE = os.path.join(WORKING_DIR, "reestr.xlsx")
RESULT_FILE = os.path.join(WORKING_DIR, "exit.xlsx")

# Column mappings for the result table
RESULT_COLUMNS = [
    "№", "ID", "код", "предмет", "кол-во", "ед.изм", "плательщик", "поставщик", 
    "пп", "дата платежа", "цена за ед.", "сумма", "отгрузка поставщиком", 
    "склад отправитель", "прием (склад получатель)", "группа", "%%", 
    "сумма перем.", "№ перемещения", "дата перемещения", "кол-во в перемещении", 
    "Соп - ие", "КОММЕНТАРИИ"
]

def load_excel_files():
    """Load the input Excel files into pandas DataFrames."""
    try:
        # First, let's read the Excel files without headers to inspect the structure
        sklad_df_raw = pd.read_excel(SKLAD_FILE, header=None, engine='openpyxl')
        reestr_df_raw = pd.read_excel(REESTR_FILE, header=None, engine='openpyxl')
        
        logger.info("Raw data structure from sklad.xlsx:")
        logger.info(sklad_df_raw.head().to_string())
        logger.info("\nRaw data structure from reestr.xlsx:")
        logger.info(reestr_df_raw.head().to_string())
        
        # Try to find the header row by looking for expected column names
        def find_header_row(df, expected_columns):
            for row_idx in range(min(10, len(df))):  # Check first 10 rows
                row_values = df.iloc[row_idx].astype(str).str.lower()
                matches = sum(1 for col in expected_columns if any(col.lower() in val for val in row_values))
                if matches >= 3:  # If we find at least 3 matches, consider it a header row
                    return row_idx
            return 0  # Default to first row if no match found
        
        # Expected column name patterns
        sklad_expected = ['код', 'наименование', 'единица', 'организация', 'номер', 'период', 'количество']
        reestr_expected = ['id', 'наименование', 'кол-во', 'компания', 'номера', 'дата', 'цена']
        
        # Find header rows
        sklad_header_row = find_header_row(sklad_df_raw, sklad_expected)
        reestr_header_row = find_header_row(reestr_df_raw, reestr_expected)
        
        logger.info(f"Detected header row in sklad.xlsx: {sklad_header_row}")
        logger.info(f"Detected header row in reestr.xlsx: {reestr_header_row}")
        
        # Read the files with the detected header rows
        sklad_df = pd.read_excel(SKLAD_FILE, header=sklad_header_row, engine='openpyxl')
        reestr_df = pd.read_excel(REESTR_FILE, header=reestr_header_row, engine='openpyxl')
        
        # Clean column names
        sklad_df.columns = sklad_df.columns.str.strip()
        reestr_df.columns = reestr_df.columns.str.strip()
        
        # Map the actual column names to expected names
        def map_columns(df, expected_patterns, is_reestr=False):
            column_mapping = {}
            
            # For reestr, we know the column indices
            if is_reestr:
                # Map known column indices for reestr (adjusting for 0-based indexing)
                column_mapping = {
                    df.columns[6]: 'ID счёта',  # Column 7
                    df.columns[5]: 'Наименование в счёте',  # Column 6
                    df.columns[17]: 'Кол-во в счёте в Ед. Изм. в счёте',  # Column 18
                    df.columns[4]: 'Компания поставщик',  # Column 5
                    df.columns[9]: 'Номера платёжных поручений',  # Column 10
                    df.columns[8]: 'Дата последнего платежа',  # Column 9
                    df.columns[21]: 'Цена за ед.'  # Column 22
                }
                
                # Log the actual column names and their mappings
                logger.info("Reestr column mappings:")
                for orig_col, new_col in column_mapping.items():
                    logger.info(f"  {orig_col} -> {new_col}")
            else:
                # For sklad, use pattern matching
                for col in df.columns:
                    col_lower = str(col).lower()
                    for pattern in expected_patterns:
                        if pattern in col_lower:
                            if 'код' in pattern and 'код' in col_lower:
                                column_mapping[col] = 'Номенклатура.Код'
                            elif 'наименование' in pattern and 'наименование' in col_lower:
                                column_mapping[col] = 'Номенклатура.Наименование'
                            elif 'единица' in pattern and 'единица' in col_lower:
                                column_mapping[col] = 'Единица'
                            elif 'организация' in pattern and 'организация' in col_lower:
                                column_mapping[col] = 'Документ связи.Организация'
                            elif 'номер' in pattern and 'номер' in col_lower:
                                column_mapping[col] = 'Регистратор.Номер'
                            elif 'период' in pattern and 'период' in col_lower:
                                column_mapping[col] = 'Период, день.Начало дня'
                            elif 'количество' in pattern and 'количество' in col_lower:
                                column_mapping[col] = 'Количество Приход'
            return column_mapping
        
        # Map columns for both dataframes
        sklad_mapping = map_columns(sklad_df, sklad_expected)
        reestr_mapping = map_columns(reestr_df, reestr_expected, is_reestr=True)
        
        logger.info("Detected column mappings for sklad.xlsx:")
        logger.info(sklad_mapping)
        logger.info("Detected column mappings for reestr.xlsx:")
        logger.info(reestr_mapping)
        
        # Rename columns
        sklad_df = sklad_df.rename(columns=sklad_mapping)
        reestr_df = reestr_df.rename(columns=reestr_mapping)
        
        # Clean data - replace NaN with empty string for string columns
        string_columns_sklad = sklad_df.select_dtypes(include=['object']).columns
        string_columns_reestr = reestr_df.select_dtypes(include=['object']).columns
        
        sklad_df[string_columns_sklad] = sklad_df[string_columns_sklad].fillna('')
        reestr_df[string_columns_reestr] = reestr_df[string_columns_reestr].fillna('')
        
        # Debug logging
        logger.info(f"Final sklad.xlsx columns: {sklad_df.columns.tolist()}")
        logger.info(f"Final reestr.xlsx columns: {reestr_df.columns.tolist()}")
        
        # Log sample data
        logger.info("Sample data from sklad.xlsx:")
        logger.info(sklad_df.head().to_string())
        logger.info("Sample data from reestr.xlsx:")
        logger.info(reestr_df.head().to_string())
        
        # Log a few rows from reestr to verify data
        logger.info("Sample rows from reestr.xlsx (Наименование в счёте):")
        for idx, row in reestr_df.head(5).iterrows():
            logger.info(f"Row {idx}: {row['Наименование в счёте']}")
        
        return sklad_df, reestr_df
    except Exception as e:
        logger.error(f"Error loading Excel files: {e}")
        raise

def get_excel_styles(file_path):
    """Extract formatting styles from the source Excel file."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header_row = ws[1]
    
    # Extract header styles
    header_font = header_row[0].font
    header_fill = header_row[0].fill
    header_border = header_row[0].border
    
    # Extract body styles
    body_row = ws[2] if ws.max_row > 1 else ws[1]
    body_font = body_row[0].font
    body_border = body_row[0].border
    
    return {
        'header_font': Font(
            name=header_font.name, size=header_font.size, 
            bold=header_font.bold, italic=header_font.italic, 
            color=header_font.color
        ),
        'header_fill': PatternFill(
            fill_type=header_fill.fill_type, 
            fgColor=header_fill.fgColor.rgb if header_fill.fgColor else None
        ),
        'header_border': Border(
            left=Side(style=header_border.left.style),
            right=Side(style=header_border.right.style),
            top=Side(style=header_border.top.style),
            bottom=Side(style=header_border.bottom.style)
        ),
        'body_font': Font(
            name=body_font.name, size=body_font.size, 
            bold=body_font.bold, italic=body_font.italic, 
            color=body_font.color
        ),
        'body_border': Border(
            left=Side(style=body_border.left.style),
            right=Side(style=body_border.right.style),
            top=Side(style=body_border.top.style),
            bottom=Side(style=body_border.bottom.style)
        )
    }

def normalize_text(text):
    text = str(text).lower()
    text = text.replace('ё', 'е')
    text = re.sub(r'[\"\'«»".,:;()\\[\\]{}<>|/\\\\\\-]', ' ', text)
    text = re.sub(r'\\s+', ' ', text)
    return text.strip()

def compare_strings_advanced(str1, str2):
    if not str1 or not str2:
        return 0

    norm1 = normalize_text(str1)
    norm2 = normalize_text(str2)

    # Точное совпадение
    if norm1 == norm2:
        return 2

    # Совпадение по подстроке
    if norm1 in norm2 or norm2 in norm1:
        return 2

    # Совпадение по ключевым словам
    set1 = set(norm1.split())
    set2 = set(norm2.split())
    intersection = set1 & set2
    union = set1 | set2
    if not union:
        return 0
    jaccard = len(intersection) / len(union)
    if jaccard > 0.6:
        return 2
    elif jaccard > 0.4:
        return 1
    else:
        return 0

def create_result_dataframe(sklad_df, reestr_df):
    """Create the result DataFrame based on the transformation algorithm."""
    result_data = []
    
    # Debug logging
    logger.info("Starting data transformation")
    logger.info(f"Number of rows in sklad: {len(sklad_df)}")
    logger.info(f"Number of rows in reestr: {len(reestr_df)}")
    
    # Step 2: Populate initial data from sklad_df
    for idx, row in sklad_df.iterrows():
        # Initialize result row with empty values
        result_row = {col: "" for col in RESULT_COLUMNS}
        
        # Copy data from sklad
        result_row["№"] = idx + 1
        result_row["код"] = row.get("Номенклатура.Код", "")
        result_row["предмет"] = row.get("Номенклатура.Наименование", "")
        result_row["ед.изм"] = row.get("Единица", "")
        result_row["плательщик"] = row.get("Документ связи.Организация", "")
        result_row["№ перемещения"] = row.get("Регистратор.Номер", "")
        result_row["дата перемещения"] = row.get("Период, день.Начало дня", "")
        result_row["кол-во в перемещении"] = row.get("Количество Приход", 0)
        result_row["группа"] = "Материалы"
        
        # Log first row data for debugging
        if idx == 0:
            logger.info(f"First row from sklad: {result_row}")
            logger.info(f"Looking for match in reestr for: {result_row['предмет']}")
        
        # Step 3: Find matching row in reestr_df
        match_level = 0
        matched_row = None
        best_similarity = 0
        
        # Get the item name from exit table (column D - "предмет")
        item_name = result_row["предмет"]
        
        # Search in reestr table column 6 ("Наименование в счёте")
        for reestr_idx, reestr_row in reestr_df.iterrows():
            reestr_name = reestr_row.get("Наименование в счёте", "")
            
            # Log comparison details for first row
            if idx == 0:
                logger.info(f"Comparing with reestr row {reestr_idx}:")
                logger.info(f"  Exit item: {item_name}")
                logger.info(f"  Reestr item: {reestr_name}")
            
            # Compare strings
            similarity = compare_strings_advanced(item_name, reestr_name)
            
            if idx == 0:
                logger.info(f"  Similarity level: {similarity}")
            
            if similarity > match_level:
                match_level = similarity
                matched_row = reestr_row
                best_similarity = similarity
        
        # Step 4: If match found, copy data from matched reestr row
        if matched_row is not None and match_level > 0:
            # Copy data from matched reestr row to result row
            result_row["ID"] = matched_row.get("ID счёта", "")
            result_row["кол-во"] = matched_row.get("Кол-во в счёте в Ед. Изм. в счёте", "")
            result_row["поставщик"] = matched_row.get("Компания поставщик", "")
            result_row["пп"] = matched_row.get("Номера платёжных поручений", "")
            result_row["дата платежа"] = matched_row.get("Дата последнего платежа", "")
            result_row["цена за ед."] = matched_row.get("Цена за ед.", "")
            
            # Calculate sum if both values are numeric
            try:
                kol_vo = float(result_row["кол-во"]) if result_row["кол-во"] else 0
                cena = float(result_row["цена за ед."]) if result_row["цена за ед."] else 0
                result_row["сумма"] = kol_vo * cena
            except (ValueError, TypeError):
                result_row["сумма"] = 0
                logger.warning(f"Could not calculate sum for row {idx + 1}")
        
        # Set match level
        result_row["Соп - ие"] = match_level
        
        # Log first row after all processing
        if idx == 0:
            logger.info(f"Final first row data: {result_row}")
        
        result_data.append(result_row)
    
    result_df = pd.DataFrame(result_data, columns=RESULT_COLUMNS)
    
    # Log summary statistics
    total_rows = len(result_df)
    matched_rows = len(result_df[result_df["Соп - ие"] > 0])
    logger.info(f"Total rows processed: {total_rows}")
    logger.info(f"Rows with matches: {matched_rows}")
    logger.info(f"Match rate: {(matched_rows/total_rows)*100:.2f}%")
    
    return result_df

def apply_excel_formatting(df, output_file, styles):
    """Save the DataFrame to Excel with formatting from the source file."""
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # Apply header formatting
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['header_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply body formatting
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = styles['body_font']
            cell.border = styles['body_border']
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_file)

def main():
    """Main function to execute the data transformation."""
    try:
        # Load data
        sklad_df, reestr_df = load_excel_files()
        
        # Get formatting styles from sklad.xlsx
        styles = get_excel_styles(SKLAD_FILE)
        
        # Create result DataFrame
        result_df = create_result_dataframe(sklad_df, reestr_df)
        
        # Save result with formatting
        apply_excel_formatting(result_df, RESULT_FILE, styles)
        
        print(f"Result file created successfully at {RESULT_FILE}")
        
    except Exception as e:
        print(f"Error: {e}")
        raise

if __name__ == "__main__":
    main()